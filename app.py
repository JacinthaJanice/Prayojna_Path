from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    send_file,
    make_response,
    flash,
    jsonify,
)
from flask_login import (
    LoginManager,
    login_user,
    logout_user,
    login_required,
    current_user,
    UserMixin,
)
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from dateutil.relativedelta import relativedelta
import json
import os
import sys
import pandas as pd
import pdfkit
import base64
import mimetypes
from io import BytesIO
import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, Alignment
import webbrowser
import threading
from jinja2 import Environment, BaseLoader

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "7777")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///users.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
DATA_FILE = "projects.json"

# Enable Jinja2 'do' extension
app.jinja_env.add_extension("jinja2.ext.do")

# Initialize SQLAlchemy
db = SQLAlchemy()
db.init_app(app)
migrate = Migrate(app, db)

# Initialize Flask-Login
login_manager = LoginManager(app)
login_manager.login_view = "login"

# Define models
user_projects = db.Table(
    "user_projects",
    db.Column("user_id", db.Integer, db.ForeignKey("user.id"), primary_key=True),
    db.Column(
        "project_no",
        db.String(50),
        db.ForeignKey("project.project_no"),
        primary_key=True,
    ),
)


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(10), nullable=False, default="user")
    projects = db.relationship(
        "Project", secondary=user_projects, backref=db.backref("users", lazy="dynamic")
    )

    def set_password(self, password):
        self.password = generate_password_hash(password, method="pbkdf2:sha256")

    def check_password(self, password):
        return check_password_hash(self.password, password)

    def get_project_access(self):
        return [project.project_no for project in self.projects]


class Project(db.Model):
    project_no = db.Column(db.String(50), primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    sanction_date = db.Column(db.Date, nullable=True)
    end_date = db.Column(db.Date, nullable=True)


@login_manager.user_loader
def load_user(user_id):
    user = db.session.get(User, int(user_id))
    logger.debug(f"Loading user ID {user_id}: {'Found' if user else 'Not found'}")
    return user


# Utility functions
def get_ordinal_suffix(n):
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def load_projects():
    data_path = get_resource_path(DATA_FILE)
    if not os.path.exists(data_path):
        logger.info(f"{data_path} not found. Initializing with empty list.")
        return []

    try:
        with open(data_path, "r") as f:
            data = json.load(f)
            for project in data:
                defaults = {
                    "status": "Ongoing",
                    "co_project_directors": [],
                    "team_members": [],  # NEW: Add team_members default
                    "pmrc_members": [],
                    "eb_members": [],
                    "pmrc_held": [],
                    "eb_held": [],
                    "cost": None,
                    "revision_cost": None,
                    "fe_cost": None,
                    "remarks": None,
                    "objectives": [],
                    "deliverables": [],
                    "attachments": [],
                    "monthly_progress": [],
                    "cost_breakdown": {},
                    "history": {
                        "end_date": [],
                        "cost": [],
                        "revision_cost": [],
                        "fe_cost": [],
                    },
                    "management_council_date": None,
                    "management_council_brief": None,
                    "management_council_mom": [],
                    "cluster_council_date": None,
                    "cluster_council_brief": None,
                    "cluster_council_mom": [],
                    "prc_pdr_date": None,
                    "prc_pdr_brief": None,
                    "prc_pdr_mom": [],
                    "prc_pdr_members": [],
                    "tiec_date": None,
                    "tiec_brief": None,
                    "tiec_mom": [],
                    "tiec_members": [],
                    "cec_date": None,
                    "cec_brief": None,
                    "cec_mom": [],
                    "dmc_date": None,
                    "dmc_brief": None,
                    "dmc_mom": [],
                    "soc_date": None,
                    "soc_brief": None,
                    "soc_mom": [],
                    "cdr_date": None,
                    "cdr_brief": None,
                    "cdr_mom": [],
                    "ddr_date": None,
                    "ddr_brief": None,
                    "ddr_mom": [],
                    "independent_committee_date": None,
                    "independent_committee_brief": None,
                    "independent_committee_mom": [],
                    "technical_closure_date": None,
                    "technical_closure_brief": None,
                    "technical_closure_mom": [],
                    "administrative_closure_date": None,
                    "administrative_closure_brief": None,
                    "administrative_closure_mom": [],
                    "closure_letter_date": None,
                    "closure_letter_brief": None,
                    "closure_letter_mom": [],
                }
                for key, value in defaults.items():
                    project.setdefault(key, value)

                if project.get("attachments"):
                    project["attachments"] = [
                        {"filename": att, "data": ""} if isinstance(att, str) else att
                        for att in project["attachments"]
                        if isinstance(att, (str, dict))
                        and (isinstance(att, str) or "filename" in att)
                    ]

                for field in [
                    "management_council_mom",
                    "cluster_council_mom",
                    "prc_pdr_mom",
                    "tiec_mom",
                    "cec_mom",
                    "dmc_mom",
                    "soc_mom",
                    "cdr_mom",
                    "ddr_mom",
                    "independent_committee_mom",
                    "technical_closure_mom",
                    "administrative_closure_mom",
                    "closure_letter_mom",
                ]:
                    if project.get(field):
                        project[field] = [
                            (
                                {"filename": item, "data": ""}
                                if isinstance(item, str)
                                else item
                            )
                            for item in project[field]
                            if isinstance(item, (str, dict))
                            and (isinstance(item, str) or "filename" in item)
                        ]

                for field in ["end_date", "cost", "revision_cost", "fe_cost"]:
                    for i, entry in enumerate(project["history"][field]):
                        entry["revision"] = (
                            "Original"
                            if i == 0
                            else get_ordinal_suffix(i) + " Revision"
                        )
            return data
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error in {data_path}: {e}")
        return []
    except Exception as e:
        logger.error(f"Error loading {data_path}: {e}")
        return []


def save_projects(projects):
    data_path = get_resource_path(DATA_FILE)
    try:
        with open(data_path, "w") as f:
            json.dump(projects, f, indent=4)
        logger.info(f"Projects saved to: {data_path}")
    except Exception as e:
        logger.error(f"Error saving projects to {data_path}: {e}")
        raise


def sync_projects_to_db():
    logger.debug("Starting project sync to database")
    projects = load_projects()
    if not projects:
        logger.info("No projects to sync")
        return

    try:
        for project in projects:
            db_project = db.session.get(Project, project["project_no"])
            if not db_project:
                try:
                    sanction_date = (
                        datetime.strptime(project["sanction_date"], "%Y-%m-%d").date()
                        if project.get("sanction_date")
                        else None
                    )
                    end_date = (
                        datetime.strptime(project["end_date"], "%Y-%m-%d").date()
                        if project.get("end_date")
                        else None
                    )
                    db_project = Project(
                        project_no=project["project_no"],
                        title=project["title"],
                        sanction_date=sanction_date,
                        end_date=end_date,
                    )
                    db.session.add(db_project)
                except ValueError as ve:
                    logger.error(
                        f"Date parsing error for project {project['project_no']}: {ve}"
                    )
                    continue
        db.session.commit()
        logger.debug("Project sync completed successfully")
    except Exception as e:
        logger.error(f"Error syncing projects to database: {e}")
        db.session.rollback()
        raise


# Initialize database and admin user
def init_db():
    with app.app_context():
        db_path = os.path.join(os.path.abspath("."), "users.db")
        logger.debug(f"Checking database at: {db_path}")

        inspector = db.inspect(db.engine)
        tables = inspector.get_table_names()
        if not all(table in tables for table in ["user", "project", "user_projects"]):
            logger.info("Creating database tables")
            db.create_all()

        admin_user = db.session.get(
            User,
            (
                User.query.filter_by(username="admin").first().id
                if User.query.filter_by(username="admin").first()
                else None
            ),
        )
        if not admin_user:
            logger.info("Creating admin user")
            admin_user = User(username="admin", role="admin")
            admin_user.set_password("admin@72$")
            db.session.add(admin_user)
            db.session.commit()
        elif not admin_user.check_password("admin@72$"):
            logger.info("Updating admin user password")
            admin_user.set_password("admin@72$")
            db.session.commit()

        sync_projects_to_db()
        logger.debug("Database initialization and project sync completed")


# Jinja2 filters
def floatformat(value, decimals=2):
    try:
        return f"{float(value):.{decimals}f}" if value is not None else "N/A"
    except (ValueError, TypeError):
        return "N/A"


app.jinja_env.filters["floatformat"] = floatformat


def format_date(date_str):
    if not date_str:
        return "N/A"
    try:
        if isinstance(date_str, datetime):
            return date_str.strftime("%d-%m-%Y")
        if "T" in date_str:
            date_obj = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
            return date_obj.strftime("%d-%m-%Y")
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        return date_obj.strftime("%d-%m-%Y")
    except ValueError:
        return date_str


app.jinja_env.filters["format_date"] = format_date


def calculate_scheduled_dates(sanction_date_str, end_date_str):
    pmrc_scheduled_dates = []
    eb_scheduled_dates = []
    if isinstance(sanction_date_str, str) and isinstance(end_date_str, str):
        try:
            sanction_date = datetime.strptime(sanction_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            current_pmrc_date = sanction_date + relativedelta(months=+3)
            while current_pmrc_date <= end_date:
                pmrc_scheduled_dates.append(current_pmrc_date.strftime("%Y-%m-%d"))
                current_pmrc_date += relativedelta(months=+3)
            current_eb_date = sanction_date + relativedelta(months=+6)
            while current_eb_date <= end_date:
                eb_scheduled_dates.append(current_eb_date.strftime("%Y-%m-%d"))
                current_eb_date += relativedelta(months=+6)
        except ValueError as e:
            logger.error(
                f"Error parsing date: sanction_date='{sanction_date_str}', end_date='{end_date_str}', error={e}"
            )
    return pmrc_scheduled_dates, eb_scheduled_dates


def generate_months(sanction_date_str, end_date_str):
    try:
        sanction_date = (
            datetime.strptime(sanction_date_str, "%Y-%m-%d")
            if sanction_date_str
            else datetime(2020, 1, 1)
        )
        end_date = (
            datetime.strptime(end_date_str, "%Y-%m-%d")
            if end_date_str
            else datetime(2025, 12, 31)
        )
    except (ValueError, TypeError) as e:
        logger.error(
            f"Date parsing error: sanction_date='{sanction_date_str}', end_date='{end_date_str}', error={e}"
        )
        sanction_date = datetime(2020, 1, 1)
        end_date = datetime(2025, 12, 31)

    months = []
    current = sanction_date.replace(day=1)
    end = end_date.replace(day=1)
    while current <= end:
        months.append(current.strftime("%Y-%m"))
        current += relativedelta(months=1)
    return months


def update_history(project, field, new_value):
    history = project["history"][field]
    timestamp = datetime.now().isoformat()
    revision_number = len(history) + 1
    revision_label = (
        "Original"
        if revision_number == 1
        else get_ordinal_suffix(revision_number - 1) + " Revision"
    )
    history.append(
        {"timestamp": timestamp, "value": new_value, "revision": revision_label}
    )
    if len(history) > 5:
        history.pop(0)
    project["history"][field] = history


# Routes
@app.route("/index")
@login_required
def index():
    logger.debug(f"Index route accessed. Current user: {current_user.username}")
    formatted_projects = []
    user_project_nos = current_user.get_project_access()
    for project in load_projects():
        if current_user.role == "admin" or project["project_no"] in user_project_nos:
            formatted_project = project.copy()
            formatted_project["sanction_date"] = format_date(
                formatted_project.get("sanction_date")
            )
            formatted_project["end_date"] = format_date(
                formatted_project.get("end_date")
            )
            formatted_projects.append(formatted_project)
    return render_template(
        "index.html", projects=formatted_projects, current_user=current_user
    )


@app.route("/")
def root():
    # Always show the login page when the site root is opened
    return render_template("login.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        logger.debug(
            f"User {current_user.username} already authenticated, redirecting to index"
        )
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        user = db.session.get(
            User,
            (
                User.query.filter_by(username=username).first().id
                if User.query.filter_by(username=username).first()
                else None
            ),
        )

        if not user:
            flash("Username not found", "error")
            logger.warning(f"Login attempt for non-existent username: {username}")
        elif user.check_password(password):
            login_user(user)
            logger.debug(f"User {username} logged in successfully")
            flash("Logged in successfully", "success")
            return redirect(url_for("index"))
        else:
            flash("Incorrect password", "error")
            logger.warning(
                f"Failed login attempt for username: {username} (incorrect password)"
            )

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logger.debug(f"User {current_user.username} logging out")
    logout_user()
    flash("Logged out successfully", "success")
    return redirect(url_for("login"))


@app.route("/manage_users", methods=["GET", "POST"])
@login_required
def manage_users():
    if current_user.role != "admin":
        flash("Access denied: Admins only.", "error")
        logger.debug(
            f"Non-admin user {current_user.username} attempted to access manage_users route"
        )
        return redirect(url_for("index"))

    users = User.query.all()
    projects = Project.query.all()

    if request.method == "POST":
        logger.debug(f"POST request received: {request.form}")
        action = request.form.get("action")

        if action == "register":
            username = request.form.get("username")
            password = request.form.get("password")
            role = request.form.get("role")
            project_access = (
                request.form.getlist("project_access") if role != "admin" else []
            )

            if not all([username, password, role]):
                logger.warning("Missing required fields")
                flash("All fields are required.", "error")
                return redirect(url_for("manage_users"))

            if len(password) < 8 or not any(c in "!@#$%^&*()" for c in password):
                logger.warning(
                    f"Password for {username} does not meet strength requirements"
                )
                flash(
                    "Password must be at least 8 characters and include a special character.",
                    "error",
                )
                return redirect(url_for("manage_users"))

            if User.query.filter_by(username=username).first():
                logger.warning(f"Username already exists: {username}")
                flash("Username already exists.", "error")
                return redirect(url_for("manage_users"))

            try:
                user = User(username=username, role=role)
                user.set_password(password)
                db.session.add(user)
                db.session.commit()

                if project_access and role != "admin":
                    db_projects = Project.query.filter(
                        Project.project_no.in_(project_access)
                    ).all()
                    user.projects.extend(db_projects)
                    db.session.commit()
                logger.info(f"User registered: {username}")
                flash("User registered successfully.", "success")
            except Exception as e:
                db.session.rollback()
                logger.error(f"Registration error for {username}: {str(e)}")
                flash(f"Registration failed: {str(e)}", "error")

        elif action in ["update", "delete"]:
            user_id = request.form.get("user_id")
            if not user_id:
                flash("User ID missing.", "error")
                return redirect(url_for("manage_users"))

            try:
                user = db.session.get(User, int(user_id))
                if not user:
                    flash("User not found.", "error")
                    return redirect(url_for("manage_users"))

                if action == "delete":
                    if user.role != "admin":
                        db.session.delete(user)
                        db.session.commit()
                        flash(f"User {user.username} deleted successfully.", "success")
                        logger.debug(
                            f"User {user.username} deleted by admin {current_user.username}"
                        )
                    else:
                        flash("Cannot delete admin users.", "error")

                elif action == "update":
                    if user.role != "admin":
                        project_access = request.form.getlist("project_access")
                        user.projects.clear()
                        if project_access:
                            db_projects = Project.query.filter(
                                Project.project_no.in_(project_access)
                            ).all()
                            if len(db_projects) != len(project_access):
                                for project_no in project_access:
                                    if not Project.query.get(project_no):
                                        json_project = next(
                                            (
                                                p
                                                for p in load_projects()
                                                if p["project_no"] == project_no
                                            ),
                                            None,
                                        )
                                        if json_project:
                                            db_project = Project(
                                                project_no=json_project["project_no"],
                                                title=json_project["title"],
                                                sanction_date=(
                                                    datetime.strptime(
                                                        json_project["sanction_date"],
                                                        "%Y-%m-%d",
                                                    ).date()
                                                    if json_project.get("sanction_date")
                                                    else None
                                                ),
                                                end_date=(
                                                    datetime.strptime(
                                                        json_project["end_date"],
                                                        "%Y-%m-%d",
                                                    ).date()
                                                    if json_project.get("end_date")
                                                    else None
                                                ),
                                            )
                                            db.session.add(db_project)
                                db.session.commit()
                                db_projects = Project.query.filter(
                                    Project.project_no.in_(project_access)
                                ).all()
                            user.projects.extend(db_projects)
                        db.session.commit()
                        flash(f"User {user.username} updated successfully.", "success")
                        logger.debug(
                            f"User {user.username} project access updated to {project_access}"
                        )
                    else:
                        flash("Cannot update admin users.", "error")

            except ValueError:
                logger.error(f"Invalid user ID: {user_id}")
                flash("Invalid user ID.", "error")
            except Exception as e:
                db.session.rollback()
                logger.error(f"Action {action} error: {str(e)}")
                flash(f"Operation failed: {str(e)}", "error")

        return redirect(url_for("manage_users"))

    return render_template("manage_users.html", users=users, projects=projects)


@app.route("/check_username", methods=["POST"])
@login_required
def check_username():
    if current_user.role != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    username = request.form.get("username")
    if not username:
        return jsonify({"available": False, "message": "Username is required"}), 400
    user = User.query.filter_by(username=username).first()
    return jsonify({"available": not bool(user)})


@app.route("/add", methods=["GET", "POST"])
@login_required
def add():
    if current_user.role != "admin":
        flash("Access denied: Admins only.", "error")
        return redirect(url_for("index"))

    if request.method == "POST":
        try:
            logger.debug(f"Processing add project form: {request.form}")
            project_no = request.form.get("project_no")
            title = request.form.get("title")
            sanction_date = request.form.get("sanction_date")
            end_date = request.form.get("end_date") or None
            project_type = request.form.get("type") or None
            project_director = request.form.get("project_director") or None
            co_project_directors = [
                d for d in request.form.getlist("co_project_directors[]") if d
            ]
            status = request.form.get("status")

            # Handle Team Members
            team_members = [
                m for m in request.form.getlist("team_members[]") if m.strip()
            ]
            if not team_members:
                team_members = []

            # Objectives with summary and status
            valid_statuses = ["Pending", "Achieved"]
            objectives_texts = request.form.getlist("objectives[]")
            objectives_statuses = [
                s if s in valid_statuses else "Pending"
                for s in request.form.getlist("objective_status[]")
            ]
            objectives = [
                {
                    "text": o,
                    "status": (
                        objectives_statuses[i]
                        if i < len(objectives_statuses)
                        else "Pending"
                    ),
                    "timestamp": datetime.now().isoformat(),
                    "summary": request.form.get(f"objective_summary_{i}", ""),
                }
                for i, o in enumerate(objectives_texts)
                if o
            ]
            if len(objectives_texts) > len(objectives_statuses):
                logger.warning(
                    f"Mismatch in objectives: {len(objectives_texts)} texts, {len(objectives_statuses)} statuses"
                )

            # Deliverables with summary and status
            deliverables_texts = request.form.getlist("deliverables[]")
            deliverables_statuses = [
                s if s in valid_statuses else "Pending"
                for s in request.form.getlist("deliverable_status[]")
            ]
            deliverables = [
                {
                    "text": d,
                    "status": (
                        deliverables_statuses[i]
                        if i < len(deliverables_statuses)
                        else "Pending"
                    ),
                    "timestamp": datetime.now().isoformat(),
                    "summary": request.form.get(f"deliverable_summary_{i}", ""),
                }
                for i, d in enumerate(deliverables_texts)
                if d
            ]
            if len(deliverables_texts) > len(deliverables_statuses):
                logger.warning(
                    f"Mismatch in deliverables: {len(deliverables_texts)} texts, {len(deliverables_statuses)} statuses"
                )

            cost = float(request.form.get("cost")) if request.form.get("cost") else None
            revision_cost = (
                float(request.form.get("revision_cost"))
                if request.form.get("revision_cost")
                else None
            )
            fe_cost = (
                float(request.form.get("fe_cost"))
                if request.form.get("fe_cost")
                else None
            )
            remarks = request.form.get("remarks") or None

            # Handle attachments
            attachments = []
            for f in request.files.getlist("attachments[]"):
                if f.filename:
                    try:
                        file_data = f.read()
                        encoded_data = base64.b64encode(file_data).decode("utf-8")
                        attachments.append(
                            {"filename": f.filename, "data": encoded_data}
                        )
                    except Exception as e:
                        flash(f"Error processing file {f.filename}: {e}", "error")

            # Handle PMRC members
            pmrc_members = []
            pmrc_indices = set()
            for key in request.form:
                if key.startswith("pmrc_members[") and "[name]" in key:
                    pmrc_indices.add(key.split("[")[1].split("]")[0])
            for index in sorted(pmrc_indices, key=int):
                name = request.form.get(f"pmrc_members[{index}][name]")
                designation = request.form.get(f"pmrc_members[{index}][designation]")
                role = request.form.get(f"pmrc_members[{index}][role]")
                if name and designation and role:
                    pmrc_members.append(
                        {"name": name, "designation": designation, "role": role}
                    )

            # Handle EB members
            eb_members = []
            eb_indices = set()
            for key in request.form:
                if key.startswith("eb_members[") and "[name]" in key:
                    eb_indices.add(key.split("[")[1].split("]")[0])
            for index in sorted(eb_indices, key=int):
                name = request.form.get(f"eb_members[{index}][name]")
                designation = request.form.get(f"eb_members[{index}][designation]")
                role = request.form.get(f"eb_members[{index}][role]")
                if name and designation and role:
                    eb_members.append(
                        {"name": name, "designation": designation, "role": role}
                    )

            # Handle monthly progress
            monthly_progress = []
            progress_indices = set()
            for key in request.form:
                if key.startswith("monthly_progress[") and "[month]" in key:
                    progress_indices.add(key.split("[")[1].split("]")[0])
            for index in sorted(progress_indices, key=int):
                month = request.form.get(f"monthly_progress[{index}][month]")
                targets = [
                    {
                        "text": t,
                        "status": (
                            request.form.getlist(
                                f"monthly_progress[{index}][target_status][]"
                            )[i]
                            if i
                            < len(
                                request.form.getlist(
                                    f"monthly_progress[{index}][target_status][]"
                                )
                            )
                            else "Pending"
                        ),
                        "timestamp": datetime.now().isoformat(),
                    }
                    for i, t in enumerate(
                        request.form.getlist(f"monthly_progress[{index}][targets][]")
                    )
                    if t
                ]
                achievements = [
                    {
                        "text": a,
                        "status": (
                            request.form.getlist(
                                f"monthly_progress[{index}][achievement_status][]"
                            )[i]
                            if i
                            < len(
                                request.form.getlist(
                                    f"monthly_progress[{index}][achievement_status][]"
                                )
                            )
                            else "Achieved"
                        ),
                        "timestamp": datetime.now().isoformat(),
                    }
                    for i, a in enumerate(
                        request.form.getlist(
                            f"monthly_progress[{index}][achievements][]"
                        )
                    )
                    if a
                ]
                if month:
                    monthly_progress.append(
                        {
                            "month": month,
                            "targets": targets,
                            "achievements": achievements,
                        }
                    )

            new_project = {
                "project_no": project_no,
                "title": title,
                "sanction_date": sanction_date,
                "end_date": end_date,
                "type": project_type,
                "project_director": project_director,
                "co_project_directors": co_project_directors,
                "team_members": team_members,
                "pmrc_members": pmrc_members,
                "eb_members": eb_members,
                "cost": cost,
                "revision_cost": revision_cost,
                "fe_cost": fe_cost,
                "pmrc_held": [],
                "eb_held": [],
                "remarks": remarks,
                "status": status,
                "objectives": objectives,
                "deliverables": deliverables,
                "attachments": attachments,
                "monthly_progress": monthly_progress,
                "cost_breakdown": {},
                "history": {
                    "end_date": [],
                    "cost": [],
                    "revision_cost": [],
                    "fe_cost": [],
                },
                "management_council_date": None,
                "management_council_brief": None,
                "management_council_mom": [],
                "cluster_council_date": None,
                "cluster_council_brief": None,
                "cluster_council_mom": [],
                "prc_pdr_date": None,
                "prc_pdr_brief": None,
                "prc_pdr_mom": [],
                "prc_pdr_members": [],
                "tiec_date": None,
                "tiec_brief": None,
                "tiec_mom": [],
                "tiec_members": [],
                "cec_date": None,
                "cec_brief": None,
                "cec_mom": [],
                "dmc_date": None,
                "dmc_brief": None,
                "dmc_mom": [],
                "soc_date": None,
                "soc_brief": None,
                "soc_mom": [],
                "cdr_date": None,
                "cdr_brief": None,
                "cdr_mom": [],
                "ddr_date": None,
                "ddr_brief": None,
                "ddr_mom": [],
                "independent_committee_date": None,
                "independent_committee_brief": None,
                "independent_committee_mom": [],
                "technical_closure_date": None,
                "technical_closure_brief": None,
                "technical_closure_mom": [],
                "administrative_closure_date": None,
                "administrative_closure_brief": None,
                "administrative_closure_mom": [],
                "closure_letter_date": None,
                "closure_letter_brief": None,
                "closure_letter_mom": [],
            }

            projects = load_projects()
            projects.append(new_project)

            # Save to database
            try:
                sanction_date_db = (
                    datetime.strptime(sanction_date, "%Y-%m-%d").date()
                    if sanction_date
                    else None
                )
                end_date_db = (
                    datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
                )
                db_project = Project(
                    project_no=project_no,
                    title=title,
                    sanction_date=sanction_date_db,
                    end_date=end_date_db,
                )
                db.session.add(db_project)
                db.session.commit()
            except ValueError as ve:
                db.session.rollback()
                logger.error(f"Date parsing error for project {project_no}: {ve}")
                flash("Error adding project to database: Invalid date format", "error")
                return redirect(url_for("add"))
            except Exception as e:
                db.session.rollback()
                logger.error(f"Error adding project {project_no} to database: {e}")
                flash(f"Error adding project to database: {e}", "error")
                return redirect(url_for("add"))

            # Save to JSON file
            try:
                save_projects(projects)
                logger.info(
                    f"Project {project_no} added with {len(objectives)} objectives, {len(deliverables)} deliverables, and {len(team_members)} team members"
                )
                flash("Project added successfully!", "success")
                return redirect(url_for("index"))
            except Exception as e:
                db.session.rollback()
                logger.error(f"Error saving projects to JSON: {e}")
                flash(f"Error saving project: {e}", "error")
                return redirect(url_for("add"))
        except Exception as e:
            logger.error(f"Error processing add project form: {e}", exc_info=True)
            flash(f"Error adding project: {e}", "error")
            return redirect(url_for("add"))

    return render_template("add.html", current_user=current_user)


@app.route("/edit/<int:index>", methods=["GET", "POST"])
@login_required
def edit(index):
    if current_user.role != "admin":
        flash("Access denied: Admins only.", "error")
        return redirect(url_for("index"))

    projects = load_projects()
    if index < 0 or index >= len(projects):
        flash("Project not found!", "error")
        return redirect(url_for("index"))

    project = projects[index]

    if request.method == "POST":
        try:
            logger.debug(f"Processing edit project form: {request.form}")
            old_end_date = project.get("end_date")
            old_cost = project.get("cost")
            old_revision_cost = project.get("revision_cost")
            old_fe_cost = project.get("fe_cost")

            project["project_no"] = request.form.get("project_no")
            project["title"] = request.form.get("title")
            project["sanction_date"] = request.form.get("sanction_date")
            project["end_date"] = request.form.get("end_date") or None
            project["type"] = request.form.get("type") or None
            project["project_director"] = request.form.get("project_director") or None
            project["co_project_directors"] = [
                d for d in request.form.getlist("co_project_directors[]") if d
            ]

            # NEW: Handle Team Members
            project["team_members"] = [
                m for m in request.form.getlist("team_members[]") if m.strip()
            ]
            if not project["team_members"]:
                project["team_members"] = []

            project["status"] = request.form.get("status")

            # Objectives with summary and status
            valid_statuses = ["Pending", "Achieved"]
            objectives_texts = request.form.getlist("objectives[]")
            objectives_statuses = [
                s if s in valid_statuses else "Pending"
                for s in request.form.getlist("objective_status[]")
            ]
            project["objectives"] = [
                {
                    "text": o,
                    "status": (
                        objectives_statuses[i]
                        if i < len(objectives_statuses)
                        else "Pending"
                    ),
                    "timestamp": datetime.now().isoformat(),
                    "summary": request.form.get(f"objective_summary_{i}", ""),
                }
                for i, o in enumerate(objectives_texts)
                if o
            ]
            if len(objectives_texts) > len(objectives_statuses):
                logger.warning(
                    f"Mismatch in objectives: {len(objectives_texts)} texts, {len(objectives_statuses)} statuses"
                )

            # Deliverables with summary and status
            deliverables_texts = request.form.getlist("deliverables[]")
            deliverables_statuses = [
                s if s in valid_statuses else "Pending"
                for s in request.form.getlist("deliverable_status[]")
            ]
            project["deliverables"] = [
                {
                    "text": d,
                    "status": (
                        deliverables_statuses[i]
                        if i < len(deliverables_statuses)
                        else "Pending"
                    ),
                    "timestamp": datetime.now().isoformat(),
                    "summary": request.form.get(f"deliverable_summary_{i}", ""),
                }
                for i, d in enumerate(deliverables_texts)
                if d
            ]
            if len(deliverables_texts) > len(deliverables_statuses):
                logger.warning(
                    f"Mismatch in deliverables: {len(deliverables_texts)} texts, {len(deliverables_statuses)} statuses"
                )

            project["cost"] = (
                float(request.form.get("cost")) if request.form.get("cost") else None
            )
            project["revision_cost"] = (
                float(request.form.get("revision_cost"))
                if request.form.get("revision_cost")
                else None
            )
            project["fe_cost"] = (
                float(request.form.get("fe_cost"))
                if request.form.get("fe_cost")
                else None
            )
            project["remarks"] = request.form.get("remarks") or None

            existing_attachments = project.get("attachments", [])
            new_attachments = []
            for f in request.files.getlist("attachments[]"):
                if f.filename:
                    try:
                        file_data = f.read()
                        encoded_data = base64.b64encode(file_data).decode("utf-8")
                        new_attachments.append(
                            {"filename": f.filename, "data": encoded_data}
                        )
                    except Exception as e:
                        flash(f"Error processing file {f.filename}: {e}", "error")

            attachments_to_keep = request.form.getlist("keep_attachments[]")
            project["attachments"] = [
                att
                for att in existing_attachments
                if att.get("filename") in attachments_to_keep
            ] + new_attachments

            project["pmrc_members"] = []
            pmrc_indices = set()
            for key in request.form:
                if key.startswith("pmrc_members[") and "[name]" in key:
                    pmrc_indices.add(key.split("[")[1].split("]")[0])
            for index in sorted(pmrc_indices, key=int):
                name = request.form.get(f"pmrc_members[{index}][name]")
                designation = request.form.get(f"pmrc_members[{index}][designation]")
                role = request.form.get(f"pmrc_members[{index}][role]")
                if name and designation and role:
                    project["pmrc_members"].append(
                        {"name": name, "designation": designation, "role": role}
                    )

            project["eb_members"] = []
            eb_indices = set()
            for key in request.form:
                if key.startswith("eb_members[") and "[name]" in key:
                    eb_indices.add(key.split("[")[1].split("]")[0])
            for index in sorted(eb_indices, key=int):
                name = request.form.get(f"eb_members[{index}][name]")
                designation = request.form.get(f"eb_members[{index}][designation]")
                role = request.form.get(f"eb_members[{index}][role]")
                if name and designation and role:
                    project["eb_members"].append(
                        {"name": name, "designation": designation, "role": role}
                    )

            project["monthly_progress"] = []
            progress_indices = set()
            for key in request.form:
                if key.startswith("monthly_progress[") and "[month]" in key:
                    progress_indices.add(key.split("[")[1].split("]")[0])
            for index in sorted(progress_indices, key=int):
                month = request.form.get(f"monthly_progress[{index}][month]")
                targets = [
                    {
                        "text": t,
                        "status": (
                            request.form.getlist(
                                f"monthly_progress[{index}][target_status][]"
                            )[i]
                            if i
                            < len(
                                request.form.getlist(
                                    f"monthly_progress[{index}][target_status][]"
                                )
                            )
                            else "Pending"
                        ),
                        "timestamp": datetime.now().isoformat(),
                    }
                    for i, t in enumerate(
                        request.form.getlist(f"monthly_progress[{index}][targets][]")
                    )
                    if t
                ]
                achievements = [
                    {
                        "text": a,
                        "status": (
                            request.form.getlist(
                                f"monthly_progress[{index}][achievement_status][]"
                            )[i]
                            if i
                            < len(
                                request.form.getlist(
                                    f"monthly_progress[{index}][achievement_status][]"
                                )
                            )
                            else "Achieved"
                        ),
                        "timestamp": datetime.now().isoformat(),
                    }
                    for i, a in enumerate(
                        request.form.getlist(
                            f"monthly_progress[{index}][achievements][]"
                        )
                    )
                    if a
                ]
                if month:
                    project["monthly_progress"].append(
                        {
                            "month": month,
                            "targets": targets,
                            "achievements": achievements,
                        }
                    )

            if project["end_date"] != old_end_date:
                update_history(project, "end_date", project["end_date"])
            if project["cost"] != old_cost:
                update_history(project, "cost", project["cost"])
            if project["revision_cost"] != old_revision_cost:
                update_history(project, "revision_cost", project["revision_cost"])
            if project["fe_cost"] != old_fe_cost:
                update_history(project, "fe_cost", project["fe_cost"])

            db_project = db.session.get(Project, project["project_no"])
            if db_project:
                db_project.title = project["title"]
                db_project.sanction_date = (
                    datetime.strptime(project["sanction_date"], "%Y-%m-%d").date()
                    if project["sanction_date"]
                    else None
                )
                db_project.end_date = (
                    datetime.strptime(project["end_date"], "%Y-%m-%d").date()
                    if project["end_date"]
                    else None
                )
            else:
                db_project = Project(
                    project_no=project["project_no"],
                    title=project["title"],
                    sanction_date=(
                        datetime.strptime(project["sanction_date"], "%Y-%m-%d").date()
                        if project["sanction_date"]
                        else None
                    ),
                    end_date=(
                        datetime.strptime(project["end_date"], "%Y-%m-%d").date()
                        if project["end_date"]
                        else None
                    ),
                )
                db.session.add(db_project)
            db.session.commit()

            save_projects(projects)
            logger.info(
                f"Project {project['project_no']} updated with {len(project['objectives'])} objectives, {len(project['deliverables'])} deliverables, and {len(project['team_members'])} team members"
            )
            flash("Project updated successfully!", "success")
            return redirect(url_for("index"))
        except Exception as e:
            db.session.rollback()
            logger.error(f"Error updating project at index {index}: {e}", exc_info=True)
            flash(f"Error updating project: {e}", "error")
            return redirect(url_for("edit", index=index))

    return render_template(
        "edit.html", project=project, index=index, current_user=current_user
    )


@app.route("/details/<int:index>")
@login_required
def details(index):
    projects = load_projects()
    if index < 0 or index >= len(projects):
        flash("Project not found!", "error")
        return redirect(url_for("index"))

    project = projects[index]
    if (
        current_user.role != "admin"
        and project["project_no"] not in current_user.get_project_access()
    ):
        flash(
            "Access denied: You do not have permission to view this project.", "error"
        )
        return redirect(url_for("index"))

    sanction_date_str = project.get("sanction_date")
    end_date_str = project.get("end_date")
    pmrc_scheduled_dates, eb_scheduled_dates = calculate_scheduled_dates(
        sanction_date_str, end_date_str
    )

    formatted_pmrc_scheduled = [format_date(date) for date in pmrc_scheduled_dates]
    formatted_eb_scheduled = [format_date(date) for date in eb_scheduled_dates]
    formatted_pmrc_held = [format_date(d) for d in project.get("pmrc_held", [])]
    formatted_eb_held = [format_date(d) for d in project.get("eb_held", [])]
    formatted_sanction_date = format_date(sanction_date_str)
    formatted_end_date = format_date(end_date_str)

    # Add analysis data
    months = generate_months(sanction_date_str, end_date_str)
    history = project.get("history", {})
    cost_changes = history.get("cost", [])
    revision_cost_changes = history.get("revision_cost", [])
    fe_cost_changes = history.get("fe_cost", [])
    end_date_changes = history.get("end_date", [])

    def format_history(entries, is_cost=True):
        return [
            {
                "revision": entry.get("revision", "N/A"),
                "value": (
                    floatformat(entry.get("value"))
                    if is_cost
                    else format_date(entry.get("value"))
                ),
                "timestamp": (
                    format_date(entry.get("timestamp").split("T")[0])
                    if entry.get("timestamp")
                    else "N/A"
                ),
            }
            for entry in entries
        ]

    formatted_cost_changes = format_history(cost_changes, is_cost=True)
    formatted_revision_cost_changes = format_history(
        revision_cost_changes, is_cost=True
    )
    formatted_fe_cost_changes = format_history(fe_cost_changes, is_cost=True)
    formatted_end_date_changes = format_history(end_date_changes, is_cost=False)

    monthly_progress = [
        {
            "month": entry.get("month", "N/A"),
            "targets": [
                {
                    "text": t.get("text", ""),
                    "status": t.get("status", "Pending"),
                    "updated_at": t.get("updated_at", t.get("timestamp", "")),
                }
                for t in entry.get("targets", [])
            ],
            "achievements": [
                {
                    "text": a.get("text", ""),
                    "status": a.get("status", "Achieved"),
                    "updated_at": a.get("updated_at", a.get("timestamp", "")),
                }
                for a in entry.get("achievements", [])
            ],
        }
        for entry in project.get("monthly_progress", [])
    ]

    objectives = [
        {
            "text": o.get("text", ""),
            "status": o.get("status", "Pending"),
            "summary": o.get("summary", ""),
            "updated_at": o.get("updated_at", o.get("timestamp", "")),
        }
        for o in project.get("objectives", [])
    ]

    deliverables = [
        {
            "text": d.get("text", ""),
            "status": d.get("status", "Pending"),
            "summary": d.get("summary", ""),
            "updated_at": d.get("updated_at", d.get("timestamp", "")),
        }
        for d in project.get("deliverables", [])
    ]

    project_for_template = project.copy()
    project_for_template["months"] = months
    project_for_template["current_month"] = datetime.now().strftime("%Y-%m")
    project_for_template["objectives"] = objectives
    project_for_template["deliverables"] = deliverables
    project_for_template["monthly_progress"] = monthly_progress
    project_for_template["cost_changes"] = formatted_cost_changes
    project_for_template["revision_cost_changes"] = formatted_revision_cost_changes
    project_for_template["fe_cost_changes"] = formatted_fe_cost_changes
    project_for_template["end_date_changes"] = formatted_end_date_changes
    project_for_template["pmrc_proposed"] = pmrc_scheduled_dates
    project_for_template["eb_proposed"] = eb_scheduled_dates
    project_for_template["pmrc_held"] = project.get("pmrc_held", [])
    project_for_template["eb_held"] = project.get("eb_held", [])
    project_for_template["attachments"] = [
        att["filename"]
        for att in project.get("attachments", [])
        if isinstance(att, dict) and "filename" in att
    ]
    active_tab = request.args.get("tab", "summary")

    return render_template(
        "details.html",
        project=project_for_template,
        pmrc_proposed=formatted_pmrc_scheduled,
        eb_proposed=formatted_eb_scheduled,
        pmrc_held=formatted_pmrc_held,
        eb_held=formatted_eb_held,
        pmrc_proposed_raw=pmrc_scheduled_dates,
        eb_proposed_raw=eb_scheduled_dates,
        pmrc_held_raw=project.get("pmrc_held", []),
        eb_held_raw=project.get("eb_held", []),
        sanction_date=formatted_sanction_date,
        end_date=formatted_end_date,
        monthly_progress=monthly_progress,
        index=index,
        active_tab=active_tab,
        current_user=current_user,
    )


@app.route("/analysis/<int:index>")
@login_required
def analysis(index):
    projects = load_projects()
    if index < 0 or index >= len(projects):
        flash("Project not found!", "error")
        return redirect(url_for("index"))

    project = projects[index]
    if (
        current_user.role != "admin"
        and project["project_no"] not in current_user.get_project_access()
    ):
        flash(
            "Access denied: You do not have permission to view this project.", "error"
        )
        return redirect(url_for("index"))

    sanction_date_str = project.get("sanction_date")
    end_date_str = project.get("end_date")
    pmrc_scheduled_dates, eb_scheduled_dates = calculate_scheduled_dates(
        sanction_date_str, end_date_str
    )

    months = generate_months(sanction_date_str, end_date_str)

    history = project.get("history", {})
    cost_changes = history.get("cost", [])
    revision_cost_changes = history.get("revision_cost", [])
    fe_cost_changes = history.get("fe_cost", [])
    end_date_changes = history.get("end_date", [])

    def format_history(entries):
        return [
            {
                "revision": entry.get("revision", "N/A"),
                "value": (
                    floatformat(entry.get("value"))
                    if "cost" in entry
                    else format_date(entry.get("value"))
                ),
                "timestamp": (
                    format_date(entry.get("timestamp").split("T")[0])
                    if entry.get("timestamp")
                    else "N/A"
                ),
            }
            for entry in entries
        ]

    formatted_cost_changes = format_history(cost_changes)
    formatted_revision_cost_changes = format_history(revision_cost_changes)
    formatted_fe_cost_changes = format_history(fe_cost_changes)
    formatted_end_date_changes = format_history(end_date_changes)

    monthly_progress = [
        {
            "month": entry.get("month", "N/A"),
            "targets": [
                {
                    "text": t.get("text", ""),
                    "status": t.get("status", "Pending"),
                    "updated_at": t.get("updated_at", t.get("timestamp", "")),
                }
                for t in entry.get("targets", [])
            ],
            "achievements": [
                {
                    "text": a.get("text", ""),
                    "status": a.get("status", "Achieved"),
                    "updated_at": a.get("updated_at", a.get("timestamp", "")),
                }
                for a in entry.get("achievements", [])
            ],
        }
        for entry in project.get("monthly_progress", [])
    ]

    objectives = [
        {
            "text": o.get("text", ""),
            "status": o.get("status", "Pending"),
            "summary": o.get("summary", ""),
            "updated_at": o.get("updated_at", o.get("timestamp", "")),
        }
        for o in project.get("objectives", [])
    ]

    deliverables = [
        {
            "text": d.get("text", ""),
            "status": d.get("status", "Pending"),
            "summary": d.get("summary", ""),
            "updated_at": d.get("updated_at", d.get("timestamp", "")),
        }
        for d in project.get("deliverables", [])
    ]

    project_for_template = project.copy()
    project_for_template["months"] = months
    project_for_template["current_month"] = datetime.now().strftime("%Y-%m")
    project_for_template["objectives"] = objectives
    project_for_template["deliverables"] = deliverables
    project_for_template["monthly_progress"] = monthly_progress
    project_for_template["cost_changes"] = formatted_cost_changes
    project_for_template["revision_cost_changes"] = formatted_revision_cost_changes
    project_for_template["fe_cost_changes"] = formatted_fe_cost_changes
    project_for_template["end_date_changes"] = formatted_end_date_changes
    # expose proposed (formerly scheduled) dates to templates
    project_for_template["pmrc_proposed"] = pmrc_scheduled_dates
    project_for_template["eb_proposed"] = eb_scheduled_dates
    project_for_template["pmrc_held"] = project.get("pmrc_held", [])
    project_for_template["eb_held"] = project.get("eb_held", [])

    return render_template(
        "analysis.html",
        project=project_for_template,
        # expose proposed (renamed) dates to templates
        pmrc_proposed=pmrc_scheduled_dates,
        eb_proposed=eb_scheduled_dates,
        pmrc_held=project.get("pmrc_held", []),
        eb_held=project.get("eb_held", []),
        monthly_progress=monthly_progress,
        index=index,
        current_user=current_user,
    )


@app.route("/update_details/<int:index>", methods=["POST"])
@login_required
def update_details(index):
    if current_user.role != "admin":
        logger.warning(f"Access denied for user {current_user.username}: not admin")
        flash("Access denied: Admins only.", "error")
        return redirect(url_for("index"))

    projects = load_projects()
    logger.debug(
        f"Received index from URL: {index}, Total projects: {len(projects)}, Projects: {[p['project_no'] for p in projects]}"
    )
    if index < 0 or index >= len(projects):
        logger.error(f"Invalid index: {index}")
        flash("Project not found!", "error")
        return redirect(url_for("index"))

    project = projects[index]
    project_index = request.form.get("project_index")
    logger.debug(
        f"Form project_index: {project_index}, Route index: {index}, Project: {project['project_no']}"
    )

    # Validate project_index
    try:
        if project_index and int(project_index) != index:
            logger.warning(
                f"Index mismatch: form project_index={project_index}, route index={index}"
            )
            flash("Invalid project index in form submission.", "warning")
            return redirect(url_for("details", index=index, tab="pre_project"))
    except (ValueError, TypeError):
        logger.error(f"Invalid project_index format: {project_index}")
        flash("Invalid project index format.", "warning")
        return redirect(url_for("details", index=index, tab="pre_project"))

    # Determine form type
    is_pre_project_form = any(
        key in request.form
        for key in [
            "management_council_date",
            "cluster_council_date",
            "prc_pdr_date",
            "tiec_date",
            "cec_date",
            "dmc_date",
            "soc_date",
            "cdr_date",
            "ddr_date",
        ]
    )
    is_project_closure_form = any(
        key in request.form
        for key in [
            "independent_committee_date",
            "technical_closure_date",
            "administrative_closure_date",
            "closure_letter_date",
        ]
    )
    is_summary_form = any(
        key.startswith("pmrc_held_")
        or key.startswith("eb_held_")
        or key.startswith("special_pmrc_")
        or key.startswith("special_eb_")
        for key in request.form
    )

    if is_summary_form:
        # Handle PMRC and EB held dates
        pmrc_held_dates = []
        eb_held_dates = []
        for key, value in request.form.items():
            if key.startswith("pmrc_held_") and value:
                pmrc_held_dates.append((int(key.split("_")[-1]), value))
            elif key.startswith("eb_held_") and value:
                eb_held_dates.append((int(key.split("_")[-1]), value))
        pmrc_held_dates.sort()
        eb_held_dates.sort()
        project["pmrc_held"] = [item[1] for item in pmrc_held_dates]
        project["eb_held"] = [item[1] for item in eb_held_dates]

        # Handle Special PMRC Meetings
        special_pmrc_meetings = []
        i = 0
        while f"special_pmrc_date_{i}" in request.form:
            date = request.form.get(f"special_pmrc_date_{i}")
            brief = request.form.get(f"special_pmrc_brief_{i}")
            if date or brief:
                special_pmrc_meetings.append({"date": date or "", "brief": brief or ""})
            i += 1
        project["special_pmrc_meetings"] = special_pmrc_meetings

        # Handle Special EB Meetings
        special_eb_meetings = []
        i = 0
        while f"special_eb_date_{i}" in request.form:
            date = request.form.get(f"special_eb_date_{i}")
            brief = request.form.get(f"special_eb_brief_{i}")
            if date or brief:
                special_eb_meetings.append({"date": date or "", "brief": brief or ""})
            i += 1
        project["special_eb_meetings"] = special_eb_meetings

    # Handle file uploads
    def handle_files(section_name):
        uploaded_files = []
        for key in request.files:
            if key.startswith(f"{section_name}_mom["):
                for f in request.files.getlist(key):
                    if f and f.filename:
                        try:
                            file_data = f.read()
                            encoded_data = base64.b64encode(file_data).decode("utf-8")
                            uploaded_files.append(
                                {"filename": f.filename, "data": encoded_data}
                            )
                        except Exception as e:
                            logger.error(f"Error processing file {f.filename}: {e}")
                            flash(f"Error processing file {f.filename}: {e}", "error")
        kept_mom = request.form.getlist(f"keep_{section_name}_mom[]")
        project[f"{section_name}_mom"] = [
            mom
            for mom in project.get(f"{section_name}_mom", [])
            if mom["filename"] in kept_mom
        ] + uploaded_files

    if is_pre_project_form:
        project["management_council_date"] = (
            request.form.get("management_council_date") or ""
        )
        project["management_council_brief"] = (
            request.form.get("management_council_brief") or ""
        )
        handle_files("management_council")

        project["cluster_council_date"] = request.form.get("cluster_council_date") or ""
        project["cluster_council_brief"] = (
            request.form.get("cluster_council_brief") or ""
        )
        handle_files("cluster_council")

        project["prc_pdr_date"] = request.form.get("prc_pdr_date") or ""
        project["prc_pdr_brief"] = request.form.get("prc_pdr_brief") or ""
        handle_files("prc_pdr")
        prc_pdr_members = []
        prc_pdr_indices = set()
        for key in request.form:
            if key.startswith("prc_pdr_members[") and "[name]" in key:
                prc_pdr_indices.add(key.split("[")[1].split("]")[0])
        for idx in sorted(prc_pdr_indices, key=int):
            name = request.form.get(f"prc_pdr_members[{idx}][name]")
            designation = request.form.get(f"prc_pdr_members[{idx}][designation]")
            if name and designation:
                prc_pdr_members.append({"name": name, "designation": designation})
        project["prc_pdr_members"] = prc_pdr_members

        project["tiec_date"] = request.form.get("tiec_date") or ""
        project["tiec_brief"] = request.form.get("tiec_brief") or ""
        handle_files("tiec")
        tiec_members = []
        tiec_indices = set()
        for key in request.form:
            if key.startswith("tiec_members[") and "[name]" in key:
                tiec_indices.add(key.split("[")[1].split("]")[0])
        for idx in sorted(tiec_indices, key=int):
            name = request.form.get(f"tiec_members[{idx}][name]")
            designation = request.form.get(f"tiec_members[{idx}][designation]")
            if name and designation:
                tiec_members.append({"name": name, "designation": designation})
        project["tiec_members"] = tiec_members

        project["cec_date"] = request.form.get("cec_date") or ""
        project["cec_brief"] = request.form.get("cec_brief") or ""
        handle_files("cec")

        project["dmc_date"] = request.form.get("dmc_date") or ""
        project["dmc_brief"] = request.form.get("dmc_brief") or ""
        handle_files("dmc")

        project["soc_date"] = request.form.get("soc_date") or ""
        project["soc_brief"] = request.form.get("soc_brief") or ""
        handle_files("soc")

        project["cdr_date"] = request.form.get("cdr_date") or ""
        project["cdr_brief"] = request.form.get("cdr_brief") or ""
        handle_files("cdr")

        project["ddr_date"] = request.form.get("ddr_date") or ""
        project["ddr_brief"] = request.form.get("ddr_brief") or ""
        handle_files("ddr")

    if is_project_closure_form:
        project["independent_committee_date"] = (
            request.form.get("independent_committee_date") or ""
        )
        project["independent_committee_brief"] = (
            request.form.get("independent_committee_brief") or ""
        )
        handle_files("independent_committee")

        project["technical_closure_date"] = (
            request.form.get("technical_closure_date") or ""
        )
        project["technical_closure_brief"] = (
            request.form.get("technical_closure_brief") or ""
        )
        handle_files("technical_closure")

        project["administrative_closure_date"] = (
            request.form.get("administrative_closure_date") or ""
        )
        project["administrative_closure_brief"] = (
            request.form.get("administrative_closure_brief") or ""
        )
        handle_files("administrative_closure")

        project["closure_letter_date"] = request.form.get("closure_letter_date") or ""
        project["closure_letter_brief"] = request.form.get("closure_letter_brief") or ""
        handle_files("closure_letter")

    try:
        save_projects(projects)
        sync_projects_to_db()
        logger.info(
            f"Project {project['project_no']} updated successfully at index {index}"
        )
        flash("Details updated successfully!", "success")
    except Exception as e:
        logger.error(f"Error saving project details: {e}", exc_info=True)
        flash(f"Error saving details: {e}", "error")
        return redirect(url_for("details", index=index, tab="pre_project"))

    redirect_tab = (
        "project_closure"
        if is_project_closure_form
        else "pre_project" if is_pre_project_form else "summary"
    )
    redirect_url = url_for("details", index=index, tab=redirect_tab)
    logger.debug(f"Redirecting to: {redirect_url}")
    response = make_response(redirect(redirect_url))
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.route("/download_attachment/<int:index>/<filename>")
@login_required
def download_attachment(index, filename):
    projects = load_projects()
    if index < 0 or index >= len(projects):
        flash("Invalid project index.", "error")
        return redirect(url_for("index"))

    project = projects[index]
    if (
        current_user.role != "admin"
        and project["project_no"] not in current_user.get_project_access()
    ):
        flash(
            "Access denied: You do not have permission to access this project.", "error"
        )
        return redirect(url_for("index"))

    sections = [
        "management_council_mom",
        "cluster_council_mom",
        "prc_pdr_mom",
        "tiec_mom",
        "cec_mom",
        "dmc_mom",
        "soc_mom",
        "cdr_mom",
        "ddr_mom",
        "independent_committee_mom",
        "technical_closure_mom",
        "administrative_closure_mom",
        "closure_letter_mom",
        "attachments",
    ]

    for section in sections:
        for mom in project.get(section, []):
            if mom["filename"] == filename:
                try:
                    if not mom.get("data"):
                        flash(
                            f'Attachment "{filename}" has no data. Please re-upload the file.',
                            "error",
                        )
                        return redirect(url_for("details", index=index))
                    file_data = base64.b64decode(mom["data"])
                    mime_type = (
                        mimetypes.guess_type(filename)[0] or "application/octet-stream"
                    )
                    return send_file(
                        BytesIO(file_data),
                        download_name=filename,
                        as_attachment=True,
                        mimetype=mime_type,
                    )
                except Exception as e:
                    flash(f"Error downloading file {filename}: {e}", "error")
                    logger.error(f"Error downloading file {filename}: {e}")
                    return redirect(url_for("details", index=index))

    flash(f"File {filename} not found.", "error")
    return redirect(url_for("details", index=index))


@app.route("/delete/<int:index>", methods=["POST"])
@login_required
def delete(index):
    if current_user.role != "admin":
        flash("Access denied: Admins only.", "error")
        return redirect(url_for("index"))

    projects = load_projects()
    if index < 0 or index >= len(projects):
        flash("Project not found!", "error")
        return redirect(url_for("index"))

    try:
        project_no = projects[index]["project_no"]
        db_project = db.session.get(Project, project_no)
        if db_project:
            db.session.delete(db_project)
            db.session.commit()
        del projects[index]
        save_projects(projects)
        flash("Project deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting project: {e}", "error")
        logger.error(f"Error deleting project at index {index}: {e}")

    return redirect(url_for("index"))


@app.route("/download_excel")
@login_required
def download_excel():
    projects = load_projects()
    if not projects:
        flash("No projects available to download", "warning")
        return redirect(url_for("index"))

    excel_data = []
    user_project_nos = current_user.get_project_access()
    for project in projects:
        if current_user.role == "admin" or project["project_no"] in user_project_nos:
            excel_data.append(
                {
                    "Project No": project.get("project_no", "N/A"),
                    "Title": project.get("title", "N/A"),
                    "Sanction Date": format_date(project.get("sanction_date", "")),
                    "End Date": format_date(project.get("end_date", "")),
                    "Type": project.get("type", "N/A"),
                    "Project Director": project.get("project_director", "N/A"),
                    "Co-Project Directors": ", ".join(
                        project.get("co_project_directors", [])
                    ),
                    "PMRC Members": ", ".join(
                        [
                            f"{m['name']} ({m['role']})"
                            for m in project.get("pmrc_members", [])
                        ]
                    ),
                    "EB Members": ", ".join(
                        [
                            f"{m['name']} ({m['role']})"
                            for m in project.get("eb_members", [])
                        ]
                    ),
                    "Cost": floatformat(project.get("cost", None)),
                    "Revision Cost": floatformat(project.get("revision_cost", None)),
                    "FE Cost": floatformat(project.get("fe_cost", None)),
                    "Remarks": project.get("remarks", "N/A"),
                    "Status": project.get("status", "N/A"),
                    "PMRC Held Dates": "\n".join(
                        [format_date(date) for date in project.get("pmrc_held", [])]
                    ),
                    "EB Held Dates": "\n".join(
                        [format_date(date) for date in project.get("eb_held", [])]
                    ),
                }
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    df = pd.DataFrame(excel_data)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        ws.append(row)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="projects_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/download_pdf")
@login_required
def download_pdf():
    projects = load_projects()
    user_project_nos = current_user.get_project_access()
    projects_data = [
        p
        for p in projects
        if current_user.role == "admin" or p["project_no"] in user_project_nos
    ]

    formatted_projects = []
    for project in projects_data:
        formatted_project = project.copy()
        formatted_project["sanction_date"] = format_date(
            formatted_project.get("sanction_date")
        )
        formatted_project["end_date"] = format_date(formatted_project.get("end_date"))
        formatted_project["cost"] = floatformat(formatted_project.get("cost"))
        formatted_project["revision_cost"] = floatformat(
            formatted_project.get("revision_cost")
        )
        formatted_project["fe_cost"] = floatformat(formatted_project.get("fe_cost"))
        formatted_project["pmrc_held"] = [
            format_date(date) for date in formatted_project.get("pmrc_held", [])
        ]
        formatted_project["eb_held"] = [
            format_date(date) for date in formatted_project.get("eb_held", [])
        ]
        pmrc_scheduled, eb_scheduled = calculate_scheduled_dates(
            project.get("sanction_date"), project.get("end_date")
        )
        formatted_project["pmrc_proposed"] = [
            format_date(date) for date in pmrc_scheduled
        ]
        formatted_project["eb_proposed"] = [format_date(date) for date in eb_scheduled]
        formatted_project["monthly_progress"] = [
            {"month": entry["month"]}
            for entry in formatted_project.get("monthly_progress", [])
        ]
        formatted_projects.append(formatted_project)

    rendered = render_template("all_projects_pdf.html", projects=formatted_projects)
    path_to_wkhtmltopdf = r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
    config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf)

    try:
        pdf = pdfkit.from_string(rendered, False, configuration=config)
        response = make_response(pdf)
        response.headers["Content-Type"] = "application/pdf"
        response.headers["Content-Disposition"] = (
            "attachment; filename=all_projects_summary.pdf"
        )
        return response
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        flash(f"Error generating PDF: {e}", "error")
        return redirect(url_for("index"))


@app.route("/download_project_pdf/<int:index>", methods=["POST"])
@login_required
def download_project_pdf(index):
    projects = load_projects()
    if index < 0 or index >= len(projects):
        flash("Project not found!", "error")
        return redirect(url_for("index"))

    project = projects[index]
    if (
        current_user.role != "admin"
        and project["project_no"] not in current_user.get_project_access()
    ):
        flash(
            "Access denied: You do not have permission to access this project.", "error"
        )
        return redirect(url_for("index"))

    sections = request.form.getlist("sections") or ["summary"]
    project_data = project.copy()
    project_data["sanction_date"] = format_date(project.get("sanction_date"))
    project_data["end_date"] = format_date(project.get("end_date"))
    project_data["cost"] = floatformat(project.get("cost"))
    project_data["revision_cost"] = floatformat(project.get("revision_cost"))
    project_data["fe_cost"] = floatformat(project.get("fe_cost"))
    project_data["pmrc_held"] = [format_date(d) for d in project.get("pmrc_held", [])]
    project_data["eb_held"] = [format_date(d) for d in project.get("eb_held", [])]
    pmrc_scheduled, eb_scheduled = calculate_scheduled_dates(
        project.get("sanction_date"), project.get("end_date")
    )
    project_data["pmrc_proposed"] = [format_date(d) for d in pmrc_scheduled]
    project_data["eb_proposed"] = [format_date(d) for d in eb_scheduled]
    project_data["total_pmrc_proposed"] = len(pmrc_scheduled)
    project_data["total_pmrc_held"] = len(project.get("pmrc_held", []))
    project_data["total_eb_proposed"] = len(eb_scheduled)
    project_data["total_eb_held"] = len(project.get("eb_held", []))
    project_data["management_council_date"] = format_date(
        project.get("management_council_date")
    )
    project_data["management_council_brief"] = project.get(
        "management_council_brief", "N/A"
    )
    project_data["cluster_council_date"] = format_date(
        project.get("cluster_council_date")
    )
    project_data["cluster_council_brief"] = project.get("cluster_council_brief", "N/A")
    project_data["prc_pdr_date"] = format_date(project.get("prc_pdr_date"))
    project_data["prc_pdr_brief"] = project.get("prc_pdr_brief", "N/A")
    project_data["tiec_date"] = format_date(project.get("tiec_date"))
    project_data["tiec_brief"] = project.get("tiec_brief", "N/A")
    project_data["cec_date"] = format_date(project.get("cec_date"))
    project_data["cec_brief"] = project.get("cec_brief", "N/A")
    project_data["dmc_date"] = format_date(project.get("dmc_date"))
    project_data["dmc_brief"] = project.get("dmc_brief", "N/A")
    project_data["soc_date"] = format_date(project.get("soc_date"))
    project_data["soc_brief"] = project.get("soc_brief", "N/A")
    project_data["cdr_date"] = format_date(project.get("cdr_date"))
    project_data["cdr_brief"] = project.get("cdr_brief", "N/A")
    project_data["ddr_date"] = format_date(project.get("ddr_date"))
    project_data["ddr_brief"] = project.get("ddr_brief", "N/A")
    project_data["independent_committee_date"] = format_date(
        project.get("independent_committee_date")
    )
    project_data["independent_committee_brief"] = project.get(
        "independent_committee_brief", "N/A"
    )
    project_data["technical_closure_date"] = format_date(
        project.get("technical_closure_date")
    )
    project_data["technical_closure_brief"] = project.get(
        "technical_closure_brief", "N/A"
    )
    project_data["administrative_closure_date"] = format_date(
        project.get("administrative_closure_date")
    )
    project_data["administrative_closure_brief"] = project.get(
        "administrative_closure_brief", "N/A"
    )
    project_data["closure_letter_date"] = format_date(
        project.get("closure_letter_date")
    )
    project_data["closure_letter_brief"] = project.get("closure_letter_brief", "N/A")

    project_data["objectives"] = [
        {
            "text": obj.get("text", ""),
            "status": obj.get("status", "N/A"),
            "summary": obj.get("summary", ""),
        }
        for obj in project.get("objectives", [])
    ]
    project_data["deliverables"] = [
        {
            "text": deliv.get("text", ""),
            "status": deliv.get("status", "N/A"),
            "summary": deliv.get("summary", ""),
        }
        for deliv in project.get("deliverables", [])
    ]
    project_data["monthly_progress"] = [
        {
            "month": entry.get("month", "N/A"),
            "targets": [
                {"text": t.get("text", ""), "status": t.get("status", "N/A")}
                for t in entry.get("targets", [])
            ],
            "achievements": [
                {"text": a.get("text", ""), "status": a.get("status", "N/A")}
                for a in entry.get("achievements", [])
            ],
        }
        for entry in project.get("monthly_progress", [])
    ]
    project_data["history"] = project.get(
        "history", {"cost": [], "revision_cost": [], "fe_cost": [], "end_date": []}
    )
    project_data["cost_breakdown"] = project.get("cost_breakdown", {})

    try:
        html_content = render_template(
            "project_pdf.html", project=project_data, sections=sections
        )
        path_to_wkhtmltopdf = r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"
        config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf)
        pdf = pdfkit.from_string(html_content, False, configuration=config)
        response = make_response(pdf)
        response.headers["Content-Type"] = "application/pdf"
        response.headers["Content-Disposition"] = (
            f'attachment; filename=project_{project.get("project_no")}_details.pdf'
        )
        return response
    except Exception as e:
        logger.error(
            f"Error generating PDF for project {project.get('project_no')}: {e}"
        )
        flash(f"Error generating PDF: {e}", "error")
        return redirect(url_for("details", index=index))


@app.route("/save_cost_breakdown/<int:index>", methods=["POST"])
@login_required
def save_cost_breakdown(index):
    if current_user.role != "admin":
        flash("Access denied: Admins only.", "error")
        return redirect(url_for("index"))

    projects = load_projects()
    if index < 0 or index >= len(projects):
        flash("Project not found!", "error")
        return redirect(url_for("index"))

    project = projects[index]
    if (
        current_user.role != "admin"
        and project["project_no"] not in current_user.get_project_access()
    ):
        flash(
            "Access denied: You do not have permission to access this project.", "error"
        )
        return redirect(url_for("index"))

    cost_breakdown = {}
    project_type = project.get("type")

    def safe_float(value):
        try:
            return float(value) if value else 0.0
        except (ValueError, TypeError):
            return 0.0

    if project_type in ["MM", "TD (T)", "TD (S)", "UT", "IF"]:
        categories = [
            "transportation",
            "equipment",
            "cars_capsi",
            "consultancy",
            "job_work",
            "hiring_transport",
            "fol_vehicles",
            "contingency",
            "plant_machinery",
            "project_vehicles",
            "works",
        ]
        total_nodal_fe = 0.0
        total_participating_fe = 0.0
        for category in categories:
            nodal_fe = safe_float(
                request.form.get(f"cost_breakdown[{category}][nodal_fe]")
            )
            participating_fe = safe_float(
                request.form.get(f"cost_breakdown[{category}][participating_fe]")
            )
            total_fe = nodal_fe + participating_fe
            cost_breakdown[category] = {
                "nodal_fe": nodal_fe,
                "participating_fe": participating_fe,
                "total_fe": total_fe,
            }
            total_nodal_fe += nodal_fe
            total_participating_fe += participating_fe

        cost_breakdown["total"] = {
            "nodal_fe": total_nodal_fe,
            "participating_fe": total_participating_fe,
            "total_fe": total_nodal_fe + total_participating_fe,
        }

    elif project_type in ["S&T (A)", "S&T (B)", "PS"]:
        revenue_categories = [
            "transportation",
            "equipment",
            "cars",
            "capsi",
            "consultancy",
            "job_work",
            "hiring_transport_fol",
            "contingency",
            "works_revenue",
        ]
        capital_categories = ["plant_machinery", "project_vehicles", "works_capital"]
        total_revenue_nodal_fe = 0.0
        total_revenue_participating_fe = 0.0
        total_capital_nodal_fe = 0.0
        total_capital_participating_fe = 0.0

        for category in revenue_categories:
            nodal_fe = safe_float(
                request.form.get(f"cost_breakdown[{category}][nodal_fe]")
            )
            participating_fe = safe_float(
                request.form.get(f"cost_breakdown[{category}][participating_fe]")
            )
            total_fe = nodal_fe + participating_fe
            cost_breakdown[category] = {
                "nodal_fe": nodal_fe,
                "participating_fe": participating_fe,
                "total_fe": total_fe,
            }
            total_revenue_nodal_fe += nodal_fe
            total_revenue_participating_fe += participating_fe

        for category in capital_categories:
            nodal_fe = safe_float(
                request.form.get(f"cost_breakdown[{category}][nodal_fe]")
            )
            participating_fe = safe_float(
                request.form.get(f"cost_breakdown[{category}][participating_fe]")
            )
            total_fe = nodal_fe + participating_fe
            cost_breakdown[category] = {
                "nodal_fe": nodal_fe,
                "participating_fe": participating_fe,
                "total_fe": total_fe,
            }
            total_capital_nodal_fe += nodal_fe
            total_capital_participating_fe += participating_fe

        cost_breakdown["total_revenue"] = {
            "nodal_fe": total_revenue_nodal_fe,
            "participating_fe": total_revenue_participating_fe,
            "total_fe": total_revenue_nodal_fe + total_revenue_participating_fe,
        }
        cost_breakdown["total_capital"] = {
            "nodal_fe": total_capital_nodal_fe,
            "participating_fe": total_capital_participating_fe,
            "total_fe": total_capital_nodal_fe + total_capital_participating_fe,
        }
        cost_breakdown["grand_total"] = {
            "nodal_fe": total_revenue_nodal_fe + total_capital_nodal_fe,
            "participating_fe": total_revenue_participating_fe
            + total_capital_participating_fe,
            "total_fe": (total_revenue_nodal_fe + total_capital_nodal_fe)
            + (total_revenue_participating_fe + total_capital_participating_fe),
        }

    project["cost_breakdown"] = cost_breakdown

    try:
        save_projects(projects)
        flash("Cost breakdown saved successfully!", "success")
    except Exception as e:
        logger.error(
            f"Error saving cost breakdown for project {project.get('project_no')}: {e}"
        )
        flash(f"Error saving cost breakdown: {e}", "error")

    response = make_response(redirect(url_for("details", index=index, tab="cost")))
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.route("/admin_login", methods=["POST"])
def admin_login():
    username = request.form.get("username")
    password = request.form.get("password")

    if username != "admin" or password != "admin@72$":
        flash("Invalid username or password.", "error")
        return redirect(url_for("login"))

    user = db.session.get(
        User,
        (
            User.query.filter_by(username="admin").first().id
            if User.query.filter_by(username="admin").first()
            else None
        ),
    )
    if not user:
        flash("Admin user not found. Please contact support.", "error")
        return redirect(url_for("login"))

    login_user(user)
    flash("Logged in as admin successfully.", "success")
    return redirect(url_for("index"))


@app.route("/download_sop_pdf")
def download_sop_pdf():
    sop_path = os.path.join(app.static_folder, "sop.pdf")
    if os.path.exists(sop_path):
        return send_file(sop_path, as_attachment=True, download_name="SOP_Document.pdf")
    else:
        flash("SOP PDF not found.", "error")
        return redirect(url_for("index"))


# Initialize database
init_db()

if __name__ == "__main__":

    def open_browser():
        webbrowser.open("http://127.0.0.1:5000")

    threading.Timer(1, open_browser).start()
    app.run(debug=True, use_reloader=False)
